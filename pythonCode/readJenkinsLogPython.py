'''
Created on March 23, 2021

@author: Srinivasulu
'''

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from time import sleep
import openpyxl 
import sys
import string
from datetime import date
import os

testPlanName = sys.argv[1]
buildURL = sys.argv[2]
# testPlanName="DAM Custom Execution"
# buildDuration="30 seconds"

# today = date.today()
# dayOfWeek = today.weekday()
print(testPlanName)
print(buildURL)

print("Execution Started")
driver=webdriver.Chrome(executable_path="C:\\wamp64\\www\\jenkinsReport\\webDriver\\chromedriver.exe")
driver.implicitly_wait(10)
driver.maximize_window()
driver.get( buildURL + "timestamps/?elapsed=HH:mm:ss.S" )

txtTimeContent=str(driver.find_element_by_tag_name("pre").text)
txtTimeElapsed=txtTimeContent.splitlines()[-1]

print(txtTimeContent)
print("Total Time: "+txtTimeElapsed)

driver.get( buildURL + "logText/progressiveText?start=0" )

txtContent=str(driver.find_element_by_tag_name("pre").text)
txtContentLastLine=txtContent.splitlines()[-1]
txtBuildStatus=txtContentLastLine.split()[-1]
print("Build Status: "+txtBuildStatus)


featureFileCount=txtContent.count("Feature file")
print(featureFileCount)


driver.close()

# filename1 = 'C:\\wamp64\\www\\jenkinsReport\\excelFiles\\testExcel.xlsx'
filename1 = str(os.getcwd())+'\\excelFiles\\testExcel.xlsx'
wb = openpyxl.load_workbook(filename1)
ws = wb.active
rowNum=ws.max_row+1

ws.cell(column=1, row=rowNum, value=testPlanName)
ws.cell(column=2, row=rowNum, value=txtTimeElapsed)
ws.cell(column=3, row=rowNum, value=txtBuildStatus)
ws.cell(column=4, row=rowNum, value=featureFileCount)
wb.save(filename1)
print("Execution Completed")
